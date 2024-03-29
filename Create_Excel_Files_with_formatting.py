#download required librairies
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border,colors
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection
from openpyxl.styles import NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl.styles.differential import DifferentialStyle

#read excel
df = pd.read_excel(r'C:\Users\mohamedirfan','Raw Data')

#add filter
df = df.loc[df['Approved/Rejected/ No Response/WIP'].isin(['Needs Action','Net New'])]
needs_action= df.loc[df['Approved/Rejected/ No Response/WIP'] == 'Needs Action']
net_new= df.loc[df['Approved/Rejected/ No Response/WIP'] == 'Net New']
Rename columns
df=df.rename(columns={'Supplier Name':'supplier_name','XSNM Qty':'Return Qty','XSNM $$':'Return $$'})

#since excel file is created with supplier name remove symbols that windows do not accept for file name
replace_symbols = ['>','<',':','/','\\\\','\|','\?','\*']
df['supplier_name'] = df['supplier_name'].replace(replace_symbols,"",regex=True)
df = df[['SupplierABK','supplier_name','Item Number','Supplier Item Number','Description','BP','Return Qty','Purchasing UOM','Return $$','Lot Number','Expiry Date','Order Number','PO Create Date']]

#Create new columns
columns= ['Status (Approved/Rejected/Other?)','Comments','Reason for Rejection or Other?','RMA#','Restocking Fee (Specify % or $)','Carrier Info','Excluded?','Approved/Rejected/ No Response/WIP']
df[columns]=''

#normalize date field
df['PO Create Date'] = pd.to_datetime(df['PO Create Date']).dt.normalize()

#groupy by supplier name
for period, data in df.groupby('supplier_name'):
    #print(data['Description'])
    output_path= os.path.join(r'C:\Irfan\python projects\Upcoming\XSNM\XSNM by Name', period +'.xlsx')
    data.to_excel(output_path,index=False)
    
    #Creating a workbook and worksheet object for formatting
    wb=openpyxl.load_workbook(output_path)
    ws=wb['Sheet1']
    
    #Insert Row for Supplier to Fill 
    ws.insert_rows(1,1)
    ws['N1'] = 'Supplier to Fill Out'
    
    #adding data validation
    data_val = DataValidation(type="list",formula1='"Approved,Rejected,Other"')
    ws.add_data_validation(data_val)
    data_val1 = DataValidation(type="list",formula1='"Too Old,Past Return Date,No Longer Carried,Discontinued"') #You can change =$A:$A with a smaller range like =A1:A9
    ws.add_data_validation(data_val1)

    data_val.promptTitle = 'Status Selection'
    data_val.prompt = 'Please select a status.'

    data_val.errorTitle = 'Invalid Status'
    data_val.error = 'Your status is not in the list of accepted entries.'

    ws.add_data_validation(data_val)

    data_val.add('N1:N1048576')


    data_val1.promptTitle = 'Reason Selection'
    data_val1.prompt = 'Please select a Reason for Rejection.'

    data_val1.errorTitle = 'Invalid Status'
    data_val1.error = 'Your status is not in the list of accepted entries.'

    ws.add_data_validation(data_val1)

    data_val1.add('P1:P1048576')
    fill_pattern = PatternFill(patternType='solid', fgColor='50E82B')
    fill_pattern1 = PatternFill(patternType='solid', fgColor='C64747')
    
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=13):
          for cell in row:
                cell.fill = fill_pattern
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=14, max_col=22):
          for cell in row:
                cell.fill = fill_pattern1

    
    #ws.conditional_formatting.add('N3:N100', CellIsRule(operator='Approved', formula=['0'],font = Font(color = '00FF0000')))
    
    #protecting worksheet            
    ws.protection.sheet = True
    #removing protection from certain columns
    for col in ['N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
        for cell in ws[col]:
            cell.protection = Protection(locked=False)
            
            
    #Changing date format
    date_style = NamedStyle(name='datetime', number_format='MM/DD/YYYY')
    date_style1 = NamedStyle(name='datetime1', number_format='MM/DD/YYYY')
    
    for rows in ws.iter_rows(min_row=3, max_row=None, min_col=13, max_col=13):
        for cell in rows:
            cell.style = date_style
            
    for rows in ws.iter_rows(min_row=3, max_row=None, min_col=11, max_col=11):
        for cell in rows:
            cell.style = date_style1
                
    #Creating spacing for cell with respect to length            
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    #saving workbook
    wb.save(output_path)

